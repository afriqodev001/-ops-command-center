"""
Read cells and full-sheet text from an uploaded .xlsx using openpyxl.

We deliberately keep this vendor-agnostic — vendor logic lives in
vendor_mappings/*. Cell keys are A1-style ("B5", "B27", etc.) and only
come from Sheet1 (or the first sheet, if it has a different name).
Sheet contents come from any named sheet other than Sheet1.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict

import openpyxl
from openpyxl.utils import get_column_letter

from .mapping_spec import ParsedPayload


_MAX_CELL_ROWS = 200
_MAX_CELL_COLS = 16


def _stringify(v) -> str:
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _read_first_sheet_cells(ws) -> Dict[str, str]:
    """Read the first N rows × M columns of a worksheet into an A1-keyed dict.

    Skips empty cells so the resulting dict is compact and easy to inspect.
    """
    out: Dict[str, str] = {}
    max_row = min(ws.max_row or 0, _MAX_CELL_ROWS)
    max_col = min(ws.max_column or 0, _MAX_CELL_COLS)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = _stringify(ws.cell(row=row, column=col).value)
            if value:
                out[f'{get_column_letter(col)}{row}'] = value
    return out


def _read_cell_labels(ws) -> Dict[str, str]:
    """For a labels-in-A / values-in-B style sheet (Epsilon's Change Details),
    return a mapping of B-cell key → the corresponding label text from column A.

    e.g. for row 5 with A5='Vendor Name' and B5='Epsilon', we return {'B5': 'Vendor Name'}.
    Empty rows are skipped.
    """
    out: Dict[str, str] = {}
    max_row = min(ws.max_row or 0, _MAX_CELL_ROWS)
    for row in range(1, max_row + 1):
        label = _stringify(ws.cell(row=row, column=1).value)
        if label:
            out[f'B{row}'] = label
    return out


def _read_sheet_text(ws) -> str:
    """Concatenate all non-empty cells of a sheet into a single text blob
    (one line per row, columns separated by tabs)."""
    rows_out = []
    max_row = min(ws.max_row or 0, _MAX_CELL_ROWS)
    max_col = min(ws.max_column or 0, _MAX_CELL_COLS)
    for row in range(1, max_row + 1):
        cells = []
        for col in range(1, max_col + 1):
            cells.append(_stringify(ws.cell(row=row, column=col).value))
        line = '\t'.join(cells).rstrip('\t')
        if line.strip():
            rows_out.append(line)
    return '\n'.join(rows_out)


def parse_xlsx(path: str | Path) -> ParsedPayload:
    """Open the xlsx and return a ParsedPayload(cells, cell_labels, sheets).

    The first sheet (whatever it's named — typically 'Change Details') is read
    as A1-keyed cells, plus a parallel `cell_labels` dict where Bn → the
    column-A label of row n. All other sheets are read as joined text.
    """
    wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            return ParsedPayload(cells={}, cell_labels={}, sheets={}, first_sheet_name='')

        first = wb[sheet_names[0]]
        cells = _read_first_sheet_cells(first)
        cell_labels = _read_cell_labels(first)

        sheets: Dict[str, str] = {}
        for name in sheet_names[1:]:
            sheets[name] = _read_sheet_text(wb[name])

        return ParsedPayload(
            cells=cells,
            cell_labels=cell_labels,
            sheets=sheets,
            first_sheet_name=sheet_names[0],
        )
    finally:
        wb.close()
